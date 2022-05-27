import openpyxl as xl
from projet_test.settings import BASE_DIR
import os
from django.db.models import Q
import array as arr
from .models import Student, Notes, Groupe, Matiere, MoyenneS1, MoyenneS2, FileModel, FilePV, Classe, Classe_Student
 
def lire_Groupes(url_file:str):
    url_absolue_str = str(BASE_DIR)
    url_absolue = url_absolue_str.replace("\\", "/")
    wkbk = xl.load_workbook(url_absolue+url_file)
    sheet = wkbk.active
    
     #Identifier la colonne de fin des matières
    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4):
        for cell in row:
            if(cell.value == "Synthèse"):
                col = cell.column

    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4, max_col = col-1):
        for cell in row:
            if(cell.value != ""):
                value = cell.value
                values = value.split(maxsplit=1)
                record = Groupe(id=values[0], libelle=values[1])
                record.save()   
    pass

def lire_Matieres(url_file:str):
    url_absolue_str = str(BASE_DIR)
    url_absolue = url_absolue_str.replace("\\", "/")
    wkbk = xl.load_workbook(url_absolue+url_file)
    sheet = wkbk.active
    classe = Classe(id=sheet['BB9'].value+'_'+sheet['BB11'].value+'_'+sheet['BB10'].value, filiere=sheet['BB10'].value, cycle=sheet['BB9'].value, niveau=sheet['BB11'].value)
    classe.save()
    
    #Identifier la colonne de fin des matières
    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4):
        for cell in row:
            if(cell.value == "Synthèse"):
                coll = cell.column
     
    #Dupliquer le noms des groupes de matières
    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4, max_col = coll-1):
        for cell in row:
            if (cell.value == ""):
                col = cell.column
                lig = cell.row 
                cell.value = sheet.cell(lig, col-1).value           

    for row in sheet.iter_rows(min_row = 16, max_row = 16, min_col = 4, max_col = coll-1):
        for cell in row:
            if((cell.value != "Synthèse UE") & (cell.value != "")):
                col = cell.column
                lig = cell.row 
                value = cell.value
                values = value.split('-')
                groupe_ide = sheet.cell(lig-1, col).value
                groupe_id = groupe_ide.split(maxsplit=1)
                record = Matiere(id=values[0], libelle=values[1], semestre = sheet['B14'].value, groupe=Groupe.objects.get(id=groupe_id[0]), classe = Classe.objects.get(id=sheet['BB9'].value+'_'+sheet['BB11'].value+'_'+sheet['BB10'].value))
                print(record)
                record.save()   
pass

def lire_student(url_file:str):
    url_absolue_str = str(BASE_DIR)
    url_absolue = url_absolue_str.replace("\\", "/")
    wkbk = xl.load_workbook(url_absolue+url_file)
    sheet = wkbk.active
    
    annee_pv=sheet['BB12'].value
    annee = annee_pv.replace("/", "-")
    
    # Identifier la ligne de fin des étudiants
    l = 0
    for row in sheet.iter_rows(min_row = 19, min_col = 2, max_col = 2):
        for cell in row:
            if(cell.value == "Effectif"):
                l = cell.row
                break
        
        if(l != 0):
            break
           
    #lire les etudiants
    for row in sheet.iter_rows(min_row = 19, max_row = l-9, min_col = 2, max_col = 2):
        for cell in row:
            lig = cell.row
            col = cell.column
            # Créer un nouvel enregistrement en utilisant la méthode d'instanciation.
            student = Student(id=cell.value, name=sheet.cell(lig, col+1).value, mot_de_passe="0000")
            student.save()
            student.classe.add(Classe.objects.get(id=sheet['BB9'].value+'_'+sheet['BB11'].value+'_'+sheet['BB10'].value))
            
            #Ajout de l'annee de la classe d'un étudiant
            classe_student = Classe_Student.objects.get(Q(student = Student.objects.get(id = cell.value)) & Q(classe = Classe.objects.get(id=sheet['BB9'].value+'_'+sheet['BB11'].value+'_'+sheet['BB10'].value)))
            classe_student.annee = annee
            classe_student.save()
            
    pass
    
    
def lire_Notes(url_file:str):
    
    url_absolue_str = str(BASE_DIR)
    url_absolue = url_absolue_str.replace("\\", "/")
    wkbk = xl.load_workbook(url_absolue+url_file)
    sheet = wkbk.active 
    
    #Identifier la colonne de fin des matières
    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4):
        for cell in row:
            if(cell.value == "Synthèse"):
                col = cell.column
    
    #Dupliquer les noms des groupes de matières   
    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4, max_col = col-1):
        for cell in row:
            if (cell.value == ""):
                col = cell.column
                lig = cell.row 
                cell.value = sheet.cell(lig, col-1).value
    
    #Dupliquer le noms des matières             
    for row in sheet.iter_rows(min_row = 16, max_row = 16, min_col = 4, max_col = col-1):
        for cell in row:
             if (cell.value == ""):
                col = cell.column
                lig = cell.row 
                cell.value = sheet.cell(lig, col-1).value
                
    # Identifier la ligne de fin des étudiants
    l = 0
    for row in sheet.iter_rows(min_row = 19, min_col = 2, max_col = 2):
        for cell in row:
            if(cell.value == "Effectif"):
                l = cell.row
                break
        
        if(l != 0):
            break
                
    cc = 0.0
    tp = 0.0
    sn = 0.0
    moyenne = 0.0
    cr = 0.0
    dec = ""
    for row in sheet.iter_rows(min_row = 19, max_row = l-9, min_col = 4, max_col = col-1):
        for cell in row:
            col = cell.column
            lig = cell.row 
            if(sheet.cell(16, col).value == sheet.cell(16, col+1).value):
                if (cell.value != ""):
                    if(sheet.cell(17, col).value == "CC"):
                        if(cell.value != None):
                            cc = cell.value
                        else:
                            cc = 0.0
                    elif(sheet.cell(17, col).value == "EXA"):
                        if(cell.value != None):
                            sn = cell.value
                        else:
                            sn = 0.0
                    elif(sheet.cell(17, col).value == "TP"):
                        if(cell.value != None):
                            tp = cell.value
                        else:
                            tp = 0.0
                    elif(sheet.cell(17, col).value == "M.Dis"):
                        if(cell.value != None):
                            moyenne = cell.value
                        else:
                            moyenne = 0.0
                    elif(sheet.cell(17, col).value == "Créd"):
                        if(cell.value != None):
                            cr = cell.value
                        else:
                            cr = 0.0
                    elif(sheet.cell(17, col).value == "Dec"):
                        if(cell.value != None):
                            dec = cell.value
                        else:
                            dec = ""
            elif(sheet.cell(16, col).value != sheet.cell(16, col+1).value):
                if (cell.value != ""):
                    if(sheet.cell(17, col).value == "CC"):
                        if(cell.value != None):
                            cc = cell.value
                        else:
                            cc = 0.0
                    elif(sheet.cell(17, col).value == "EXA"):
                        if(cell.value != None):
                            sn = cell.value
                        else:
                            sn = 0.0
                    elif(sheet.cell(17, col).value == "TP"):
                        if(cell.value != None):
                            tp = cell.value
                        else:
                            tp = 0.0
                    elif(sheet.cell(17, col).value == "M.Dis"):
                        if(cell.value != None):
                            moyenne = cell.value
                        else:
                            moyenne = 0.0
                    elif(sheet.cell(17, col).value == "Créd"):
                        if(cell.value != None):
                            cr = cell.value
                        else:
                            cr = 0.0
                    elif(sheet.cell(17, col).value == "Dec"):
                        if(cell.value != None):
                            dec = cell.value
                        else:
                            dec = ""
                
                if((sheet.cell(16, col).value != "Synthèse UE") & (moyenne != 0.0) & (moyenne != None)):   
                   value = sheet.cell(16, col).value
                   values = value.split('-')
                   record = Notes(id = sheet.cell(lig, 2).value+'_'+sheet.cell(16, col).value, note_cc = cc, note_tp = tp, note_sn = sn, note_moy = moyenne, credit = cr, decision = dec, student = Student.objects.get(id=sheet.cell(lig, 2).value), matiere = Matiere.objects.get(id=values[0]))
                   record.save()
                   cc = 0.0
                   tp = 0.0
                   sn = 0.0
                   moyenne = 0.0
                   cr = 0.0
                   dec = ""
                elif(sheet.cell(16, col).value == "Synthèse UE"):
                    cc = 0.0
                    tp = 0.0
                    sn = 0.0
                    moyenne = 0.0
                    cr = 0.0
                    dec = ""
    pass

def lire_Moy(url_file:str):
    url_absolue_str = str(BASE_DIR)
    url_absolue = url_absolue_str.replace("\\", "/")
    wkbk = xl.load_workbook(url_absolue+url_file)
    sheet = wkbk.active
    
    #Identifier la colonne de fin des moyennes générales
    for row in sheet.iter_rows(min_row = 15, max_row = 15, min_col = 4):
        for cell in row:
            if(cell.value == "Synthèse"):
                col = cell.column
                
   # Identifier la ligne de fin des étudiants
    l = 0
    for row in sheet.iter_rows(min_row = 19, min_col = 2, max_col = 2):
        for cell in row:
            if(cell.value == "Effectif"):
                l = cell.row
                break
        
        if(l != 0):
            break
    
    for row in sheet.iter_rows(min_row = 19, max_row = l-9, min_col = col, max_col = col):
        for cell in row:
            colonne = cell.column
            ligne = cell.row
            annee_pv = sheet['BB12'].value
            annee = annee_pv.replace("/", "-")
            if(sheet['B14'].value == "SEMESTRE 1"):
                record = MoyenneS1(id = "Moyenne_Semestre_1_"+sheet.cell(ligne, 2).value+'_'+sheet['BB12'].value, annee = annee, moy_s1 = cell.value, mgp = sheet.cell(ligne, colonne+1).value, credit = sheet.cell(ligne, colonne+3).value, credit_total = sheet.cell(18, colonne+4).value, student=Student.objects.get(id=sheet.cell(ligne, 2).value))
                record.save()
            elif(sheet['B14'].value == "SEMESTRE 2"):
                record = MoyenneS2(id = "Moyenne_Semestre_2_"+sheet.cell(ligne, 2).value+'_'+sheet['BB12'].value, annee = annee, moy_s2 = cell.value, mgp = sheet.cell(ligne, colonne+1).value, credit = sheet.cell(ligne, colonne+3).value, credit_total = sheet.cell(18, colonne+4).value, student=Student.objects.get(id=sheet.cell(ligne, 2).value))
                record.save()
    
    pass
