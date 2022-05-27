import openpyxl as xl
from openpyxl.styles import PatternFill, NamedStyle, Color
# from openpyxl.styles.colors import RED
from projet_test.settings import BASE_DIR
import os
import mimetypes
from django.http.response import HttpResponse
from django.db.models import Q
import array as arr
from .models import Student, Notes, Groupe, Matiere, MoyenneS1, MoyenneS2, FileModel, FilePV, Classe, Classe_Student, User



#Demander à Fabrice si il est nécessaire de nommer le fichier
def generate_excel(excel_file_path, name):
    
    url_absolue_str = str(BASE_DIR)
    url_absolue = url_absolue_str.replace("\\", "/")  
    
    # To get file from server
    filepath = url_absolue+excel_file_path
    path = open(file = filepath, mode = 'rb',)
    
    # To get the type of the filer
    mime_type, _ = mimetypes.guess_type(filepath)
    
    # To give a name to the file that'll be dowbloaded
    # new_file_name = name
    print(name)
    
    # To generate excel file and auto download
    response = HttpResponse(path, content_type = mime_type)
    response['Content-Disposition'] = "attachment; filename=%s" % name
    return response

def Students_Classe(classe, annee):
    students_classe = Classe_Student.objects.filter(Q(classe = classe) & Q(annee = annee))
    return students_classe

def Matieres_Classe(classe, semestre):
    matieres_classe = Matiere.objects.filter(Q(classe = classe) & Q(semestre = semestre))   
    return matieres_classe

def rattrapages_list(classe, semestre, annee):
    liste = []
    students_classe = Students_Classe(classe, annee)
    matieres = Matieres_Classe(classe, semestre)
    for i in students_classe:
        for j in matieres:
           note = Notes.objects.filter((Q(decision="EL") | Q(decision="NC") | Q(decision="CANT")) & Q(matiere = j) & Q(student = i.student))
           if(note.exists()):
               liste.append(note)
                     
    return(liste)
    
   

def excel_rattrapage(liste, semestre, classe, annee):
    
    students_classe = Students_Classe(classe, annee)
    matieres = Matieres_Classe(classe, semestre)
      
    model = FileModel.objects.get(name = "Rattrapages")  
    print(model.file_model)   
    url_absolue_str = str(BASE_DIR)
    url_file = str(model.file_model)
    url_absolue = url_absolue_str.replace("\\", "/")  
    wkbk = xl.load_workbook(url_absolue+'/'+url_file)
    sheet = wkbk.active 
    
    sheet['E6'].value = semestre
    sheet['E7'].value = classe.cycle
    sheet['E8'].value = classe.filiere
    sheet['E9'].value = classe.niveau
    sheet['E10'].value = annee
    
    #Lister les étudiants dans le excel
    row = 15
    for s in students_classe:
        sheet.cell(row, 3).value = s.student.name
        row = row+1
    
    #Lister les matieres dans le excel
    col = 4
    for m in matieres:
        sheet.cell(13, col).value = m.libelle
        col = col+1 
    
    #Sélection des cellules correspondantes aux rattrapages
    for row in sheet.iter_rows(min_row = 15, max_row = 15+len(students_classe), min_col = 4, max_col = 4+len(matieres)):
        for cell in row:
            lig = cell.row
            col = cell.column
            student = sheet.cell(lig, 3)
            matiere = sheet.cell(13, col)
            for i in liste:
                for j in i:
                    if((sheet.cell(lig, 3).value == j.student.name) & (sheet.cell(13, col).value == j.matiere.libelle)):
                        if(j.decision == "CANT"):
                            # cell.value = "CANT"
                            cell.fill = xl.styles.PatternFill(patternType = 'solid', fgColor = 'FFFF00')
                        else:
                            cell.fill = xl.styles.PatternFill(patternType = 'solid', fgColor = 'FF0000')
                                                
                        
    #Génération du fichier excel
    # annee_url = annee.replace("/", "-")       
    wkbk.save(url_absolue+'/media/Rattrapages_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx')
    excel_file_path = '/media/Rattrapages_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx'
    name = 'Rattrapages_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx'
    
    url_name = {
        "url" : excel_file_path,
        "name" : name
    }
    # return model2
    return url_name

 
def recap_notes_students(classe_student, semestre):
    
    liste_notes = []
    matieres = Matieres_Classe(classe, semestre)
    for j in matieres:
        note = Notes.objects.filter(Q(matiere = j) & Q(student = classe_student.student))
        if(note.exists()):
            liste_notes.append(note) 
    
    return liste_notes 


def excel_bilan_annuel_classe(classe, annee):
    
    liste = []
    liste_studentsS1  = []
    liste_studentsS2  = []
    liste_students = []
    infos_bilanS1 = []
    infos_bilanS2 = []
    students = Students_Classe(classe, annee)
    print(students)
    
    for student in students: 
        print(student.annee)
    
    for student in students: 
        infos_bilanS1.append(MoyenneS1.objects.filter(Q(annee = student.annee) & Q(student = student.student)))
   
    for student in students: 
        infos_bilanS2.append(MoyenneS2.objects.filter(Q(annee = student.annee) & Q(student = student.student)))

    # for i in infos_bilanS1:
    #    for j in i:
    #        print(j.mgp) 
           
    for i in infos_bilanS1:
        for j in i:
            liste_studentsS1.append(j.student)
    
    for i in infos_bilanS2:
        for j in i:
            liste_studentsS2.append(j.student)
    
    for stud in  liste_studentsS1:
        if(stud in liste_studentsS2):
            liste_students.append(stud) 
    
    print(liste_students)
           
    for student in liste_students:
        for infoS1 in infos_bilanS1:
            for info1 in infoS1:
                if(info1.student == student):
                    # print(info1.mgp)
                    mgp1 = info1.mgp
                    credit1 = info1.credit
                    break
        for infoS2 in infos_bilanS2:
            for info2 in infoS2:
                if(info2.student == student):
                    # print(info2.mgp)
                    mgp2 = info2.mgp
                    credit2 = info2.credit
                    break
                
        print("Annuel") 
        mgp = (mgp1 + mgp2)/2 
        credit =  credit1 + credit2
        print(mgp1, " plus ", mgp2, " egal ", mgp)  
        print(credit1, " plus ", credit2, " egal ", credit)  
        info_annuel_student = {
            "student_id" : student.id,
            "student_name" : student.name,
            "credit" : credit,
            "mgp" : mgp
        }
        liste.append(info_annuel_student)
    
    #Classer par ordre décroissant
    liste.sort(key=lambda x: x.get('mgp'), reverse=True)
    for l in liste:
        print(l.get("student_id"))
        # for valeur in l:
        #     print(valeur.get(student_id))
            
    model = FileModel.objects.get(name = "BilanAnnuel")  
    print(model.file_model)   
    url_absolue_str = str(BASE_DIR)
    url_file = str(model.file_model)
    url_absolue = url_absolue_str.replace("\\", "/")  
    wkbk = xl.load_workbook(url_absolue+'/'+url_file)
    sheet = wkbk.active 
    
    sheet['E7'].value = classe.cycle
    sheet['E8'].value = classe.filiere
    sheet['E9'].value = classe.niveau
    sheet['E10'].value = annee
    
    row = 16
    col = 1
    i = 1
    for l in liste:
        print(l.get("student_id"))
        sheet.cell(row, col).value = i
        sheet.cell(row, col+1).value = l.get("student_id")
        sheet.cell(row, col+2).value = l.get("student_name")
        sheet.cell(row, col+3).value = sheet['E8'].value
        i = i+1
        row = row + 1
    
    c = 0
    credit_annuel = 0
    coll = col+3
    column = col+3   
    if(len(infos_bilanS1) != 0):
        c = 2
        # sheet.cell(14, column).value = "Bilan Semestre 1" 
        sheet.cell(14, column+1).value = "Bilan Semestre 1"
        sheet.merge_cells(start_row = 14, start_column = column+1, end_row = 14, end_column = column+2)
        sheet.cell(15, column+1).value = "MGP"
        sheet.cell(15, column+2).value = "CREDIT"
        lig = 16
        for l in liste:
            for infoS1 in infos_bilanS1:
                for i in infoS1:
                    if(l.get("student_id") == i.student.id):
                        if(i.mgp < 2.00):
                            sheet.cell(lig, column+1).font = xl.styles.Font(color = 'FF0000')
                            sheet.cell(lig, column+1).value = i.mgp
                        else:
                            sheet.cell(lig, column+1).font = xl.styles.Font(color = '00FF00')
                            sheet.cell(lig, column+1).value = i.mgp
                            
                        if(i.credit < i.credit_total):
                            sheet.cell(lig, column+2).font = xl.styles.Font(color = 'FF0000')
                            sheet.cell(lig, column+2).value = i.credit
                        else:
                            sheet.cell(lig, column+2).font = xl.styles.Font(color = '00FF00')
                            sheet.cell(lig, column+2).value = i.credit
                            
                        lig = lig+1
                        credit_total1 = i.credit_total
                        break
    
    column = column + c   
    if(len(infos_bilanS2) != 0):
        c = 2
        # sheet.cell(14, column).value = "Bilan Semestre 1" 
        sheet.cell(14, column+1).value = "Bilan Semestre 2"
        sheet.merge_cells(start_row = 14, start_column = column+1, end_row = 14, end_column = column+2)
        sheet.cell(15, column+1).value = "MGP"
        sheet.cell(15, column+2).value = "CREDIT"
        
        lig = 16
        for l in liste:
            for infoS2 in infos_bilanS2:
                for i in infoS2:
                    if(l.get("student_id") == i.student.id):
                        if(i.mgp < 2.00):
                            sheet.cell(lig, column+1).font = xl.styles.Font(color = 'FF0000')
                            sheet.cell(lig, column+1).value = i.mgp
                        else:
                            sheet.cell(lig, column+1).font = xl.styles.Font(color = '00FF00')
                            sheet.cell(lig, column+1).value = i.mgp
                            
                        if(i.credit < i.credit_total):
                            sheet.cell(lig, column+2).font = xl.styles.Font(color = 'FF0000')
                            sheet.cell(lig, column+2).value = i.credit
                        else:
                            sheet.cell(lig, column+2).font = xl.styles.Font(color = '00FF00')
                            sheet.cell(lig, column+2).value = i.credit
                            
                        lig = lig+1
                        credit_total2 = i.credit_total
                        break
    
    column = column + c
    if(len(liste) != 0):
        sheet.cell(13, coll+1).value = "Niveau "+classe.niveau       
        sheet.merge_cells(start_row = 13, start_column = coll+1, end_row = 13, end_column = column)  
        
        sheet.cell(13, column+1).value = "Bilan Niveau "+classe.niveau
        sheet.merge_cells(start_row = 13, start_column = column+1, end_row = 14, end_column = column+4)
        
        sheet.cell(15, column+1).value = "MGP"
        sheet.cell(15, column+2).value = "CREDIT"
        sheet.cell(15, column+3).value = "DECISION"
        sheet.cell(15, column+4).value = "RANG"
        
        lig = 16
        i = 1
        credit_annuel = int(credit_total1) + int(credit_total2)
        for l in liste:
            # sheet.cell(lig, column+1).value = l.get("mgp")
            # sheet.cell(lig, column+2).value = l.get("credit")
            if(l.get("mgp") < 2.00):
                sheet.cell(lig, column+1).font = xl.styles.Font(color = 'FF0000')
                sheet.cell(lig, column+1).value = l.get("mgp")
            else:
                sheet.cell(lig, column+1).font = xl.styles.Font(color = '00FF00')
                sheet.cell(lig, column+1).value = l.get("mgp")
                            
            if(l.get("credit") < credit_annuel):
                sheet.cell(lig, column+2).font = xl.styles.Font(color = 'FF0000')
                sheet.cell(lig, column+2).value = l.get("credit")
            else:
                sheet.cell(lig, column+2).font = xl.styles.Font(color = '00FF00')
                sheet.cell(lig, column+2).value = l.get("credit")
                
            if(l.get("mgp") >= 2.00 ):
                niveau = int(classe.niveau) + int(1)
                sheet.cell(lig, column+3).value = "Admis en "+str(niveau)+"e Annee"
                sheet.cell(lig, column+3).font = xl.styles.Font(color = '00FF00')
            else:
                sheet.cell(lig, column+3).value = "Echec"
                sheet.cell(lig, column+3).font = xl.styles.Font(color = 'FF0000')
                
            sheet.cell(lig, column+4).value = i
            lig = lig+1
            i = i+1
    
    
    #Génération du fichier excel           
    wkbk.save(url_absolue+'/media/BilanAnnuel_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx')
    excel_file_path = '/media/BilanAnnuel_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx'
    name = 'BilanAnnuel_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx'
    # model2 = FileRattrapage(id='Rattrapages_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee_url+'.xlsx', annee = annee, file_rattrapage = 'media/Rattrapages_'+classe.cycle+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx')
    # model2.save() 
    # return generate_excel(excel_file_path, name) 
    url_name = {
        "url" : excel_file_path,
        "name" : name,
    }
    # return model2
    return url_name  
      
    pass


def niveau_cycle(classe):
    
    niveau_cycle = []
    niveaux = Classe.objects.filter(cycle = classe.cycle)
    
    for niveau in niveaux:
        if(niveau.niveau <= classe.niveau):
            # niveau_cycle.append(niveau.niveau)
            if(classe.cycle == "Ingenieur"):
                if(classe.niveau >= '3'):
                    if(niveau.niveau > '2'):
                        niveau_cycle.append(niveau)
                    # niveau_cycle.append(niveau)
                else:
                    niveau_cycle.append(niveau)
            else:
                niveau_cycle.append(niveau)
            
    # if(classe.cycle == "Ingenieur"):
    #     if(classe.niveau <= 2 ):
    #         niveau_cycle.append(niveau.niveau)
    
    # print(niveau_cycle)
    return niveau_cycle
    
    pass

def bilan_cycle(classe, annee):
    
    annee_cycle = annee.split('-')
    
    print(annee_cycle)
    
    classe_cycle = niveau_cycle(classe)
    for c in classe_cycle:
        print(c.niveau)
    
    # print(classe_cycle)
    
    annees = []
    listes = []
    
    
    nbre_de_classe = len(classe_cycle)
    
    print(nbre_de_classe)
    
    for i in range(nbre_de_classe):
        annee_debut = int(annee_cycle[0]) - i
        annee_fin = int(annee_cycle[1]) - i
        annee = str(annee_debut)+'-'+str(annee_fin)
        annees.append(annee)
    
    annees.sort()
    
    print(annees)
    
    for i in range(len(classe_cycle)):
        
        liste = []
        liste_studentsS1  = []
        liste_studentsS2  = []
        liste_students = []
        infos_bilanS1 = []
        infos_bilanS2 = []
        liste_infos_annuel = []
        classe = classe_cycle[i]
        annee = annees[i]
        
        
        # students = Classe_Student.objects.filter(Q(classe = classe_cycle[i]) & Q(annee = annees[i]))
        students = Classe_Student.objects.filter(Q(classe = classe) & Q(annee = annee))
       
        # print(students)
        
        # for i in students:
        #     print(i.student)
    
        # for student in students: 
        #     print(student.annee)
    
        for student in students: 
            infos_bilanS1.append(MoyenneS1.objects.filter(Q(annee = student.annee) & Q(student = student.student)))
        # print(infos_bilanS1)            
        liste_infos_annuel.append(infos_bilanS1)
        
        for student in students: 
            infos_bilanS2.append(MoyenneS2.objects.filter(Q(annee = student.annee) & Q(student = student.student)))
        
        liste_infos_annuel.append(infos_bilanS2)
           
        for i in infos_bilanS1:
            for j in i:
                liste_studentsS1.append(j.student)
    
        for i in infos_bilanS2:
            for j in i:
                liste_studentsS2.append(j.student)
    
        for stud in  liste_studentsS1:
            if(stud in liste_studentsS2):
                liste_students.append(stud) 
    
        # print(liste_students)
        
        for student in liste_students:
            for infoS1 in infos_bilanS1:
                for info1 in infoS1:
                    if(info1.student == student):
                        # print(info1.mgp)
                        mgp1 = info1.mgp
                        credit1 = info1.credit
                        break
            for infoS2 in infos_bilanS2:
                for info2 in infoS2:
                    if(info2.student == student):
                        # print(info2.mgp)
                        mgp2 = info2.mgp
                        credit2 = info2.credit
                        break
                
            # print("Annuel") 
            mgp = (mgp1 + mgp2)/2 
            credit =  credit1 + credit2
            # print(mgp1, " plus ", mgp2, " egal ", mgp)  
            # print(credit1, " plus ", credit2, " egal ", credit)  
            info_annuel_student = {
                "student_id" : student.id,
                "student_name" : student.name,
                "credit" : credit,
                "mgp" : mgp,
                "niveau_classe" : classe.niveau
            }
            liste.append(info_annuel_student)
    
        #Classer par ordre décroissant
        liste.sort(key=lambda x: x.get('mgp'), reverse=True)
        liste_infos_annuel.append(liste)
        listes.append(liste_infos_annuel)
            
    print(len(listes))
    
    i = 0
    j = 0
    sum_mgp = 0
    sum_moy = 0
    n = len(listes)
    liste_finale = []
    print(len(listes))
    while i < n:
        j = 0
        m = len(listes[i][2])
        if(len(liste_finale) == 0):
            verif = "NO"
        while j < m:
            if(len(liste_finale) == 0):
                info_annuel_student = {
                    "student_id" : listes[i][2][j].get("student_id"),
                    "student_name" : listes[i][2][j].get("student_name"),
                    "credit" : listes[i][2][j].get("credit"),
                    "mgp" : listes[i][2][j].get("mgp")
                }
                liste_finale.append(info_annuel_student)
                # sum_mgp = sum_mgp + listes[i][2][j].get("mgp")
                # print(listes[i][2][j].get("student_id")+"  ", sum_mgp)
                verif == "OK"
                j = j + 1
            else:
                stop = 1
                for l in liste_finale:
                    if(l.get("student_id") == listes[i][2][j].get("student_id")):
                        l["credit"] = l.get("credit") + listes[i][2][j].get("credit")
                        l["mgp"] = l.get("mgp") + listes[i][2][j].get("mgp")
                        stop = 0
                        break
                    
                if(stop == 1):
                    info_annuel_student = {
                        "student_id" : listes[i][2][j].get("student_id"),
                        "student_name" : listes[i][2][j].get("student_name"),
                        "credit" : listes[i][2][j].get("credit"),
                        "mgp" : listes[i][2][j].get("mgp")
                    }
                    liste_finale.append(info_annuel_student)
                    
                j = j + 1 
        i = i+1 
    
    liste_finale.sort(key=lambda x: x.get('mgp'), reverse=True)
    # print(type(liste_finale[0]))   
    
    
    model = FileModel.objects.get(name = "BilanCycle")  
    print(model.file_model)   
    url_absolue_str = str(BASE_DIR)
    url_file = str(model.file_model)
    url_absolue = url_absolue_str.replace("\\", "/")  
    wkbk = xl.load_workbook(url_absolue+'/'+url_file)
    sheet = wkbk.active 
    
    sheet['E7'].value = classe.cycle
    sheet['E8'].value = classe.filiere
    sheet['E9'].value = classe.niveau
    sheet['E10'].value = annee
       
                    
    row = 18
    col = 1
    i = 1
    for l in liste_finale:
        print(l.get("student_id"))
        sheet.cell(row, col).value = i
        sheet.cell(row, col+1).value = l.get("student_id")
        sheet.cell(row, col+2).value = l.get("student_name")
        i = i+1
        row = row + 1
    
    
    column = col+2
    c = 0  
     
    if(len(listes[0]) != 0):
        fin = column
        for liste_annuel in listes:
            sum_c = 0
            if(len(liste_annuel) != 0):
                # debut = col+3
                for liste in liste_annuel:
                    column = column+c
                    ligne = 18
            
                    if(len(liste) != 0):
                        c = 3
                        sum_c = sum_c+c 
                        
                        if(liste == liste_annuel[0]):
                            sheet.cell(16, column+1).value = "Bilan Semestre 1"
                            sheet.merge_cells(start_row = 16, start_column = column+1, end_row = 16, end_column = column+3)
                        elif(liste == liste_annuel[1]):
                            sheet.cell(16, column+1).value = "Bilan Semestre 2"
                            sheet.merge_cells(start_row = 16, start_column = column+1, end_row = 16, end_column = column+3)
                        elif(liste == liste_annuel[2]):
                            sheet.cell(16, column+1).value = "Bilan Annuel"
                            sheet.merge_cells(start_row = 16, start_column = column+1, end_row = 16, end_column = column+3)
                            
                            
                        sheet.cell(17, column+1).value = "MGP"
                        sheet.cell(17, column+2).value = "CREDIT"
                        sheet.cell(17, column+3).value = "GRADE"
                        
                        for liste_infos in liste_finale:                                
                                
                            for l in liste:
                                if(isinstance(l, dict)):
                                    if(l.get("student_id") == liste_infos.get("student_id")):
                                        
                                        niveau = l.get("niveau_classe")
                                        
                                        if(l.get("mgp") < 2.00):
                                            sheet.cell(ligne, column+1).value = l.get("mgp")
                                            sheet.cell(ligne, column+1).font = xl.styles.Font(color = 'FF0000')
                                        else:
                                            sheet.cell(ligne, column+1).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+1).value = l.get("mgp")
                                        
                                        sheet.cell(ligne, column+2).value = l.get("credit")
                                        
                                        if(3.80 <= float(l.get("mgp")) <= 4.0):
                                            sheet.cell(ligne, column+3).value = "A+"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(3.40 <= float(l.get("mgp")) < 3.80):
                                            sheet.cell(ligne, column+3).value = "A"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(3.10 <= float(l.get("mgp")) < 3.40):
                                            sheet.cell(ligne, column+3).value = "B+"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(2.80 <= float(l.get("mgp")) < 3.10):
                                            sheet.cell(ligne, column+3).value = "B"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(2.40 <= float(l.get("mgp")) < 2.80):
                                            sheet.cell(ligne, column+3).value = "B-"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(2.10 <= float(l.get("mgp")) < 2.40):
                                            sheet.cell(ligne, column+3).value = "C+"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(2.0 <= float(l.get("mgp")) < 2.10):
                                            sheet.cell(ligne, column+3).value = "C"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                        elif(1.80 <= float(l.get("mgp")) < 2.0):
                                            sheet.cell(ligne, column+3).value = "C"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                        elif(1.40 <= float(l.get("mgp")) < 1.80):
                                            sheet.cell(ligne, column+3).value = "C-"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                        elif(1.10 <= float(l.get("mgp")) < 1.40):
                                            sheet.cell(ligne, column+3).value = "D"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                        elif(0.10 <= float(l.get("mgp")) < 1.10):
                                            sheet.cell(ligne, column+3).value = "E"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                        elif(float(l.get("mgp")) < 0.1):
                                            sheet.cell(ligne, column+3).value = "F"
                                            sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                            sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                        else:
                                            sheet.cell(ligne, column+3).value = l.get("mgp")
                                        
                                        break
                                else:
                                    for n in l:
                                        if(n.student.id == liste_infos.get("student_id")):
                                            
                                            if(n.mgp < 2.00):
                                                sheet.cell(ligne, column+1).value = n.mgp
                                                sheet.cell(ligne, column+1).font = xl.styles.Font(color = 'FF0000')
                                            else:
                                                sheet.cell(ligne, column+1).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+1).value = n.mgp
                                            
                                            sheet.cell(ligne, column+2).value = n.credit
                                            
                                            if(3.80 <= float(n.mgp) <= 4.0):
                                                sheet.cell(ligne, column+3).value = "A+"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(3.40 <= float(n.mgp) < 3.80):
                                                sheet.cell(ligne, column+3).value = "A"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(3.10 <= float(n.mgp) < 3.40):
                                                sheet.cell(ligne, column+3).value = "B+"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(2.80 <= float(n.mgp) < 3.10):
                                                sheet.cell(ligne, column+3).value = "B"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(2.40 <= float(n.mgp) < 2.80):
                                                sheet.cell(ligne, column+3).value = "B-"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(2.10 <= float(n.mgp) < 2.40):
                                                sheet.cell(ligne, column+3).value = "C+"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(2.0 <= float(n.mgp) <= 2.099):
                                                sheet.cell(ligne, column+3).value = "C"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                                            elif(1.80 <= float(n.mgp) < 2.0):
                                                sheet.cell(ligne, column+3).value = "C"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                            elif(1.40 <= float(n.mgp) < 1.80):
                                                sheet.cell(ligne, column+3).value = "C-"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                            elif(1.1 <= float(n.mgp) < 1.40):
                                                sheet.cell(ligne, column+3).value = "D"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                            elif(0.1 <= float(n.mgp) < 1.1):
                                                sheet.cell(ligne, column+3).value = "E"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                            elif(float(n.mgp) < 0.1):
                                                sheet.cell(ligne, column+3).value = "F"
                                                sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                                                sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                                            else:
                                                sheet.cell(ligne, column+3).value = n.mgp
                                            
                                            break
                                      
                            ligne = ligne+1
                fin = column+3
                debut = fin - sum_c
                sheet.cell(15, debut+1).value = "Niveau "+str(niveau)
                sheet.merge_cells(start_row = 15, start_column = debut+1, end_row = 15, end_column = fin)
        
        
        column = fin
        ligne = 18
        
        sheet.cell(15, column+1).value = "BILAN CYCLE"
        sheet.merge_cells(start_row = 15, start_column = column+1, end_row = 16, end_column = column+5)
        sheet.cell(17, column+1).value = "MGP"
        sheet.cell(17, column+2).value = "CREDIT"
        sheet.cell(17, column+3).value = "GRADE"
        sheet.cell(17, column+4).value = "DECISION"
        sheet.cell(17, column+5).value = "RANG"
        rang = 1
        for l in liste_finale:
            
            if(isinstance(l, dict)):
                
                if((float(l.get("mgp"))/2) < 2.00):
                    sheet.cell(ligne, column+1).value = (float(l.get("mgp"))/2)
                    sheet.cell(ligne, column+1).font = xl.styles.Font(color = 'FF0000')
                else:
                    sheet.cell(ligne, column+1).font = xl.styles.Font(color = '00FF00')
                    sheet.cell(ligne, column+1).value = (float(l.get("mgp"))/2)
                
                sheet.cell(ligne, column+5).value = rang
                
                sheet.cell(ligne, column+2).value = l.get("credit")
                
                if(3.8 <= (float(l.get("mgp"))/2) <= 4.0):
                    sheet.cell(ligne, column+3).value = "A+"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(3.4 <= (float(l.get("mgp"))/2) < 3.8):
                    sheet.cell(ligne, column+3).value = "A"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(3.1 <= (float(l.get("mgp"))/2) < 3.4):
                    sheet.cell(ligne, column+3).value = "B+"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(2.8 <= (float(l.get("mgp"))/2) < 3.1):
                    sheet.cell(ligne, column+3).value = "B"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(2.4 <= (float(l.get("mgp"))/2) < 2.8):
                    sheet.cell(ligne, column+3).value = "B-"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(2.1 <= (float(l.get("mgp"))/2) < 2.4):
                    sheet.cell(ligne, column+3).value = "C+"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(2.0 <= (float(l.get("mgp"))/2) < 2.1):
                    sheet.cell(ligne, column+3).value = "C"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+4).value = "Admis"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = '00FF00')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = '00FF00')
                elif(1.8 <= (float(l.get("mgp"))/2) < 2.0):
                    sheet.cell(ligne, column+3).value = "C"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+4).value = "Echec"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                elif(1.4 <= (float(l.get("mgp"))/2) < 1.8):
                    sheet.cell(ligne, column+3).value = "C-"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+4).value = "Echec"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                elif(1.1 <= (float(l.get("mgp"))/2) < 1.4):
                    sheet.cell(ligne, column+3).value = "D"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+4).value = "Echec"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                elif(0.1 <= (float(l.get("mgp"))/2) < 1.1):
                    sheet.cell(ligne, column+3).value = "E"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+4).value = "Echec"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                elif((float(l.get("mgp"))/2) < 0.1):
                    sheet.cell(ligne, column+3).value = "F"
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+4).value = "Echec"
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')
                else:
                    sheet.cell(ligne, column+3).value = l.get("mgp")
                    sheet.cell(ligne, column+3).font = xl.styles.Font(color = 'FF0000')
                    
                    sheet.cell(ligne, column+4).font = xl.styles.Font(color = 'FF0000')
                    sheet.cell(ligne, column+4).value = "Error"
                    
                    sheet.cell(ligne, column+2).font = xl.styles.Font(color = 'FF0000')  
                          
            ligne = ligne+1
            rang = rang+1
            
            
    #Génération du fichier excel           
    wkbk.save(url_absolue+'/media/BilanCycle_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx')
    excel_file_path = '/media/BilanCycle_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx'
    name = 'BilanCycle_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx'
    # model2 = FileRattrapage(id='Rattrapages_'+classe.cycle+'_'+classe.niveau+'_'+classe.filiere+'_'+annee_url+'.xlsx', annee = annee, file_rattrapage = 'media/Rattrapages_'+classe.cycle+classe.niveau+'_'+classe.filiere+'_'+annee+'.xlsx')
    # model2.save() 
    # return generate_excel(excel_file_path, name) 
    url_name = {
        "url" : excel_file_path,
        "name" : name,
    }
    # return model2
    return url_name  
            
    
    

    
   
    pass     
